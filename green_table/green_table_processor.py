import os
import glob
import json
import re
import pandas as pd
from openpyxl import load_workbook
import traceback
from datetime import datetime

class GreenTableProcessor:
    def __init__(self, log_callback=None):
        self.log_callback = log_callback
        self.dictionary = {}
        self.shikhta_templates = {}
        self.error_log = []
        
        # Load dictionary and prepare patterns immediately
        self.load_dictionary()

    def log_message(self, message):
        if self.log_callback:
            self.log_callback(message)
        else:
            print(message)

    def load_dictionary(self):
        """Загрузка JSON-словаря из файла"""
        default_dict = {
            "litera_plavki": ["А", "АБ", "АВ", "АГ", "АЕ", "Б", "В", "Г", "Д", "Е", "Ж", 
                             "И", "К", "Л", "М", "Н", "П", "Р", "С", "Т", "ТЭ", "У", "Ф", 
                             "Ц", "Ш", "Щ", "Ю", "Я", "ГО", "ВО", "C"],
            "metally": ["GTD-111", "GTD-111 DS", "Mar M247", "ВЖЛ14-ВИ", "ЖС26-ВИ", 
                       "ЖС26У-ВИ", "ЖС6К-ВИ", "ЗМИ-3У-ВИ", "Сталь 21-11-2,5", "DS200+Hf", 
                       "CMSX-4", "Inconel 100", "Inconel 939", "ЧС70У-ВИ", "ВЖМ4-ВИ",
                       "ВЖМ-200", "Ni-W", "CMSX-4", "ВЖМ 4 ВИ (24ЖР-200И)", "24CMSX-4-168B",
                       " GTD11 В24-248", "GTD11", "РИ", "Mar", "ВЖМ", "Inconel", "GTD"],
            "shikhta_patterns": [
                "23И-467Н", "200010503", "23ВН-487Н", "23И-57Н", "РИ61803",
                "24-И-4033", "24Н-267Н", "25-Н-4003", "25-И4005", "В22-134",
                "В24Ж1-105В", "В24Ж1-103В", "23Ж1-155В", "РИ 65108", "25-Н-4001"
            ],
            "kommentarii_etalony": []
        }
        
        try:
            # Получаем путь к директории, где находится скрипт
            program_dir = os.path.dirname(os.path.abspath(__file__))
            dict_path = os.path.join(program_dir, 'dictionary.json')
            
            if os.path.exists(dict_path):
                with open(dict_path, 'r', encoding='utf-8') as f:
                    self.dictionary = json.load(f)
                self.log_message(f"Словарь успешно загружен из файла {dict_path}")
            else:
                self.dictionary = default_dict
                with open(dict_path, 'w', encoding='utf-8') as f:
                    json.dump(self.dictionary, f, ensure_ascii=False, indent=2)
                self.log_message(f"Создан файл {dict_path} со словарем по умолчанию")
        except Exception as e:
            self.log_message(f"Ошибка при загрузке словаря: {str(e)}")
            self.dictionary = default_dict
        
        # Подготавливаем паттерны шихты
        self.prepare_shikhta_patterns()

    def prepare_shikhta_patterns(self):
        """Подготовка паттернов шихты"""
        self.shikhta_templates = {}
        
        for pattern in self.dictionary.get('shikhta_patterns', []):
            try:
                pattern_str = str(pattern)
                # Очищаем паттерн от специальных символов
                clean_pattern = re.sub(r'[^A-Za-zА-Яа-я0-9]', '', pattern_str)
                
                if not clean_pattern:
                    continue
                
                # Преобразуем в шаблон LNLNLN...
                template_parts = []
                for char in clean_pattern:
                    if char.isdigit():
                        template_parts.append('N')
                    elif char.isalpha():
                        template_parts.append('L')
                    else:
                        template_parts.append('?')
                
                template = ''.join(template_parts)
                
                # Добавляем в словарь шаблонов
                if template not in self.shikhta_templates:
                    self.shikhta_templates[template] = {
                        'original_patterns': [],
                        'regexes': []
                    }
                
                self.shikhta_templates[template]['original_patterns'].append(pattern)
                
                # Создаем regex для поиска этого паттерна
                regex_pattern = ''
                for i, char in enumerate(clean_pattern):
                    if i > 0:
                        regex_pattern += r'[^A-Za-zА-Яа-я0-9]*'
                    
                    if char.isdigit():
                        regex_pattern += r'\d'
                    elif char.isalpha():
                        regex_pattern += f'[{char.lower()}{char.upper()}]'
                
                self.shikhta_templates[template]['regexes'].append(
                    re.compile(regex_pattern, re.IGNORECASE)
                )
                
            except Exception as e:
                self.log_message(f"Ошибка при обработке паттерна {pattern}: {str(e)}")
        
        self.log_message(f"Подготовлено {len(self.shikhta_templates)} уникальных шаблонов шихты")

    def damerau_levenshtein_distance(self, s1, s2):
        """Расстояние Дамерау-Левенштейна"""
        if not s1 or not s2:
            return max(len(s1 or ''), len(s2 or ''))
        
        s1 = str(s1)
        s2 = str(s2)
        
        len1 = len(s1)
        len2 = len(s2)
        
        d = [[0] * (len2 + 1) for _ in range(len1 + 1)]
        
        for i in range(len1 + 1):
            d[i][0] = i
        for j in range(len2 + 1):
            d[0][j] = j
        
        for i in range(1, len1 + 1):
            for j in range(1, len2 + 1):
                cost = 0 if s1[i-1] == s2[j-1] else 1
                
                d[i][j] = min(
                    d[i-1][j] + 1,
                    d[i][j-1] + 1,
                    d[i-1][j-1] + cost
                )
                
                if (i > 1 and j > 1 and s1[i-1] == s2[j-2] and s1[i-2] == s2[j-1]):
                    d[i][j] = min(d[i][j], d[i-2][j-2] + cost)
        
        return d[len1][len2]

    def get_cell_value(self, sheet, merged_ranges, row, col):
        """Получает значение ячейки с учетом объединенных ячеек"""
        for merged_range in merged_ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and 
                merged_range.min_col <= col <= merged_range.max_col):
                return sheet.cell(merged_range.min_row, merged_range.min_col).value
        return sheet.cell(row, col).value

    # --- RECOGNITION METHODS ---

    def is_plavka(self, value):
        if not value: return None
        value_str = str(value).strip()
        literas_pattern = '|'.join(self.dictionary['litera_plavki'])
        pattern = r'(\d{2})(' + literas_pattern + r')(\d{1,3})'
        match = re.search(pattern, value_str, re.IGNORECASE)
        if match:
            year, litera, number = match.groups()
            return f"{year}{litera}{number}"
        pattern2 = r'(\d{2})([А-Я]{1,2})(\d{1,3})'
        match2 = re.search(pattern2, value_str, re.IGNORECASE)
        if match2:
            year, litera, number = match2.groups()
            if litera.upper() in self.dictionary['litera_plavki']:
                return f"{year}{litera.upper()}{number}"
        return None

    def is_metall(self, value):
        if not value: return None
        value_str = str(value).strip()
        if value_str in self.dictionary['metally']:
            return value_str
        
        best_match = None
        min_distance = float('inf')
        for metall in self.dictionary['metally']:
            distance = self.damerau_levenshtein_distance(value_str, metall)
            max_len = max(len(value_str), len(metall))
            normalized_distance = distance / max_len if max_len > 0 else 1
            if normalized_distance <= 0.5 and distance < min_distance:
                min_distance = distance
                best_match = value_str
        return best_match

    def is_shikhta(self, value):
        if not value: return None
        value_str = str(value)
        parts = re.split(r'[\s,;:/\\|]+', value_str)
        
        for part in parts:
            if not part: continue
            clean_part = re.sub(r'[^A-Za-zА-Яа-я0-9]', '', part)
            if not clean_part or len(clean_part) < 3: continue
            
            for pattern in self.dictionary.get('shikhta_patterns', []):
                clean_pattern = re.sub(r'[^A-Za-zА-Яа-я0-9]', '', str(pattern))
                if not clean_pattern: continue
                if clean_part == clean_pattern: return value_str
                if clean_pattern in clean_part: return value_str
                if clean_part in clean_pattern: return value_str
                
                regex_pattern = ''
                for i, char in enumerate(clean_pattern):
                    if i > 0: regex_pattern += r'[^A-Za-zА-Яа-я0-9]*'
                    if char.isdigit(): regex_pattern += r'\d'
                    elif char.isalpha(): 
                        if char.islower(): regex_pattern += f'[{char}{char.upper()}]'
                        else: regex_pattern += f'[{char}{char.lower()}]'
                try:
                    regex = re.compile(regex_pattern, re.IGNORECASE)
                    if regex.search(clean_part): return value_str
                except re.error: continue

            part_template_parts = []
            for char in clean_part:
                if char.isdigit(): part_template_parts.append('N')
                elif char.isalpha(): part_template_parts.append('L')
                else: part_template_parts.append('?')
            part_template = ''.join(part_template_parts)
            
            for template, template_data in self.shikhta_templates.items():
                if template == part_template or template in part_template or part_template in template:
                    for regex in template_data['regexes']:
                        if regex.search(clean_part): return value_str
        return None

    def is_zero_or_one(self, value):
        if value is None: return None
        try:
            if isinstance(value, (int, float)):
                num = int(value)
                if num in [0, 1] and float(value) == num: return num
            else:
                value_str = str(value).strip()
                if value_str in ['0', '1']: return int(value_str)
        except (ValueError, TypeError): pass
        return None

    def is_blok(self, value):
        if value is None: return None
        zero_or_one = self.is_zero_or_one(value)
        if zero_or_one is not None: return None
        try:
            if isinstance(value, (int, float)):
                num = int(value)
                if 0 <= num <= 9999 and float(value) == num: return num
            else:
                value_str = str(value).strip()
                if value_str.isdigit():
                    num = int(value_str)
                    if 0 <= num <= 9999: return num
        except (ValueError, TypeError): pass
        return None

    def is_obrazec(self, value):
        if not value: return None
        value_str = str(value).strip().lower()
        if 'обр' in value_str or 'затравк' in value_str: return str(value)
        return None

    def extract_litera_from_plavka(self, plavka_value):
        if not plavka_value: return None
        plavka_str = str(plavka_value)
        match = re.search(r'\d{2}([А-Я]{1,2})\d{1,3}', plavka_str, re.IGNORECASE)
        if match: return match.group(1).upper()
        return None

    def is_nomenklatura(self, value, plavka_litera):
        if not value: return None
        value_str = str(value).strip()
        if plavka_litera in ['А', 'Г'] and '522' in value_str: return str(value)
        if 'иятл' in value_str.lower(): return str(value)
        return None

    def is_tu(self, value):
        if not value: return None
        value_str = str(value).strip()
        if re.search(r'ту', value_str, re.IGNORECASE): return str(value)
        return None

    # --- STRUCTURE ANALYSIS ---

    def find_large_merged_cells(self, sheet, merged_ranges):
        large_cells = []
        for merged_range in merged_ranges:
            height = merged_range.max_row - merged_range.min_row + 1
            width = merged_range.max_col - merged_range.min_col + 1
            value = sheet.cell(merged_range.min_row, merged_range.min_col).value
            value_str = str(value or '').lower()
            is_large = (height > 3 or width > 8) or 'перебор' in value_str
            if is_large:
                large_cells.append({
                    'range': merged_range,
                    'min_row': merged_range.min_row, 'max_row': merged_range.max_row,
                    'min_col': merged_range.min_col, 'max_col': merged_range.max_col,
                    'height': height, 'width': width, 'value': value,
                    'is_perebor': 'перебор' in value_str
                })
        return large_cells

    def check_group_for_perebor(self, group_start_row, col_start, large_cells):
        group_end_row = group_start_row + 2
        group_end_col = col_start + 7
        for cell_info in large_cells:
            row_overlap = (group_start_row <= cell_info['max_row'] and group_end_row >= cell_info['min_row'])
            col_overlap = (col_start <= cell_info['max_col'] and group_end_col >= cell_info['min_col'])
            if row_overlap and col_overlap and cell_info['is_perebor']:
                return True, cell_info
        return False, None

    def check_group_structure(self, sheet, merged_ranges, group_start_row, col_start):
        ceramic_valid = False
        iron_valid = False
        for merged_range in merged_ranges:
            if (merged_range.min_row == group_start_row and merged_range.min_col == col_start and
                merged_range.max_row == group_start_row + 2 and merged_range.max_col == col_start):
                ceramic_valid = True
                break
        for merged_range in merged_ranges:
            if (merged_range.min_row == group_start_row and merged_range.min_col == col_start + 1 and
                merged_range.max_row == group_start_row + 2 and merged_range.max_col == col_start + 1):
                iron_valid = True
                break
        return ceramic_valid, iron_valid

    def find_chain_ranges(self, sheet, merged_ranges):
        empty_rows = []
        for row_idx in range(1, sheet.max_row + 1):
            is_empty = True
            has_fill = False
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = self.get_cell_value(sheet, merged_ranges, row_idx, col_idx)
                if cell_value is not None and str(cell_value).strip() != '':
                    is_empty = False; break
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.fill.start_color.index != '00000000':
                    has_fill = True; break
            if is_empty and not has_fill: empty_rows.append(row_idx)
        
        chain_ranges = []
        start_row = 1
        for empty_row in empty_rows:
            if empty_row > start_row: chain_ranges.append((start_row, empty_row - 1))
            start_row = empty_row + 1
        if start_row <= sheet.max_row: chain_ranges.append((start_row, sheet.max_row))
        return chain_ranges

    def find_furnace_for_non_uvnk(self, sheet, merged_ranges, start_row, col_start, col_end, data_start_row, data_end_row):
        found_uppf = False
        furnace_value = ""
        for r in range(start_row, start_row + 3):
            for c in range(col_start, col_end + 1):
                cell_val = self.get_cell_value(sheet, merged_ranges, r, c)
                if cell_val and "УППФ" in str(cell_val):
                    furnace_value = cell_val
                    found_uppf = True
                    break
            if found_uppf: break
        if not found_uppf:
            for r in range(data_start_row, data_end_row + 1):
                for c in range(col_start, col_end + 1):
                    cell_val = self.get_cell_value(sheet, merged_ranges, r, c)
                    if cell_val and "УППФ" in str(cell_val):
                        furnace_value = cell_val
                        found_uppf = True
                        break
                if found_uppf: break
        return furnace_value if found_uppf else ""

    # --- PROCESSING ---

    def process_group(self, sheet, merged_ranges, group_start_row, col_start, 
                    date_value, furnace_value, is_uvnk, processed_merge_keys,
                    sheet_name, chain_idx, block_idx, group_idx):
        
        ceramic_valid, iron_valid = self.check_group_structure(sheet, merged_ranges, group_start_row, col_start)
        is_standard_group = ceramic_valid and iron_valid
        
        if not is_standard_group:
            self.log_message(f"    Группа {group_idx+1}: Нестандартная структура - обработка как Spec.Group")
        
        group_data = {
            'Дата': date_value, 'Печь': furnace_value, 'Керамика': None, 'Железо': None,
            'Плавка': None, 'ТУ': None, 'Металл': None, 'Шихта': None,
            'Блок 1': None, 'Блок 2': None, 'Комплект оснастки': None,
            'Образцы': None, 'Номенклатура': None, 'Комментарий': []
        }
        
        contains_uppf = False
        processed_cells = set()
        values_to_process = []
        
        for row_offset in range(3):
            row = group_start_row + row_offset
            for col_offset in range(8):
                col = col_start + col_offset
                cell_value = self.get_cell_value(sheet, merged_ranges, row, col)
                if cell_value in (None, '', ' '): continue
                if cell_value and "УППФ" in str(cell_value): contains_uppf = True
                
                is_merged_cell = False
                merge_key = None
                for merged_range in merged_ranges:
                    if (merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col):
                        is_merged_cell = True
                        merge_key = (merged_range.min_row, merged_range.min_col)
                        break
                
                if is_merged_cell and merge_key in processed_cells: continue
                if is_merged_cell: processed_cells.add(merge_key)
                else: processed_cells.add((row, col))
                
                values_to_process.append({'value': cell_value, 'row_offset': row_offset, 'col_offset': col_offset})
        
        if contains_uppf:
            self.log_message(f"    Группа {group_idx+1}: Пропущена (содержит УППФ)")
            return None
        
        has_data = any(item['value'] not in (None, '', ' ') for item in values_to_process)
        if not has_data and not group_data['Комментарий']:
            self.log_message(f"    Группа {group_idx+1}: Пропущена (пустая)")
            return None

        if is_standard_group:
            for item in values_to_process:
                if item['row_offset'] == 0 and item['col_offset'] == 0:
                    group_data['Керамика'] = item['value']; item['processed'] = True; break
            for item in values_to_process:
                if item['row_offset'] == 0 and item['col_offset'] == 1:
                    group_data['Железо'] = item['value']; item['processed'] = True; break
            
            remaining = [item for item in values_to_process if not item.get('processed')]
            plavka_litera = None
            
            # Sequence: Plavka, TU, Metall, Shikhta, Blok, 0/1, Obrazec, Nomenklatura
            for item in remaining:
                if item.get('processed'): continue
                plavka = self.is_plavka(item['value'])
                if plavka and not group_data['Плавка']:
                    group_data['Плавка'] = plavka
                    plavka_litera = self.extract_litera_from_plavka(plavka)
                    item['processed'] = True
            
            for item in remaining:
                if item.get('processed'): continue
                tu_val = self.is_tu(item['value'])
                if tu_val and not group_data['ТУ']:
                    group_data['ТУ'] = tu_val; item['processed'] = True; break
            
            for item in remaining:
                if item.get('processed'): continue
                metall = self.is_metall(item['value'])
                if metall and not group_data['Металл']:
                    group_data['Металл'] = metall; item['processed'] = True
            
            for item in remaining:
                if item.get('processed'): continue
                shikhta = self.is_shikhta(item['value'])
                if shikhta and not group_data['Шихта']:
                    group_data['Шихта'] = shikhta; item['processed'] = True

            for item in remaining:
                if item.get('processed'): continue
                blok = self.is_blok(item['value'])
                if blok is not None:
                    if is_uvnk:
                        if group_data['Блок 1'] is None: group_data['Блок 1'] = blok; item['processed'] = True
                        elif group_data['Блок 2'] is None: group_data['Блок 2'] = blok; item['processed'] = True
                    else:
                        if group_data['Блок 1'] is None: group_data['Блок 1'] = blok; item['processed'] = True

            for item in remaining:
                if item.get('processed'): continue
                zo = self.is_zero_or_one(item['value'])
                if zo is not None:
                    if group_data['Блок 1'] is not None and not group_data['Комплект оснастки']: 
                         group_data['Комплект оснастки'] = zo; item['processed'] = True
                    elif group_data['Блок 1'] is None:
                         group_data['Блок 1'] = zo; item['processed'] = True
            
            for item in remaining:
                 if item.get('processed'): continue
                 obrazec = self.is_obrazec(item['value'])
                 if obrazec is not None and not group_data['Образцы']:
                      group_data['Образцы'] = obrazec; item['processed'] = True

            for item in remaining:
                 if item.get('processed'): continue
                 if plavka_litera:
                      nom = self.is_nomenklatura(item['value'], plavka_litera)
                      if nom and not group_data['Номенклатура']:
                           group_data['Номенклатура'] = nom; item['processed'] = True
            
            for item in remaining:
                 if not item.get('processed'): group_data['Комментарий'].append(str(item['value']))

        else:
            plavka_litera = None
            for item in values_to_process:
                val = item['value']
                plavka = self.is_plavka(val)
                if plavka:
                     if not group_data['Плавка']:
                          group_data['Плавка'] = plavka
                          plavka_litera = self.extract_litera_from_plavka(plavka)
                
                shikhta = self.is_shikhta(val)
                if shikhta and not group_data['Шихта']: group_data['Шихта'] = shikhta
                
                metall = self.is_metall(val)
                if metall and not group_data['Металл']: group_data['Металл'] = metall
                
                tu_val = self.is_tu(val)
                if tu_val and not group_data['ТУ']: group_data['ТУ'] = tu_val
                
                obrazec = self.is_obrazec(val)
                if obrazec and not group_data['Образцы']: group_data['Образцы'] = obrazec

                blok = self.is_blok(val)
                if blok is not None:
                     if is_uvnk:
                          if group_data['Блок 1'] is None: group_data['Блок 1'] = blok
                          elif group_data['Блок 2'] is None and group_data['Блок 1'] != blok: group_data['Блок 2'] = blok
                     else:
                          if group_data['Блок 1'] is None: group_data['Блок 1'] = blok
                
                if plavka_litera:
                     nom = self.is_nomenklatura(val, plavka_litera)
                     if nom and not group_data['Номенклатура']: group_data['Номенклатура'] = nom
                
                is_recognized = (plavka or shikhta or metall or tu_val or obrazec or 
                                 (blok is not None) or (plavka_litera and self.is_nomenklatura(val, plavka_litera)))
                if not is_recognized: group_data['Комментарий'].append(str(val))

        if group_data['Комментарий']:
             group_data['Комментарий'] = '; '.join(list(set(group_data['Комментарий'])))
        else:
             group_data['Комментарий'] = ''
             
        return group_data

    def process_sheet(self, workbook, sheet_name):
        sheet = workbook[sheet_name]
        self.log_message(f"Обработка листа: {sheet_name}")
        has_uvnk = 'увнк' in sheet_name.lower()
        
        furnace_value = ""
        if has_uvnk:
            sheet_name_lower = sheet_name.lower()
            if '1' in sheet_name_lower[sheet_name_lower.find('увнк'):]: furnace_value = 'УВНК-9А1'
            elif '2' in sheet_name_lower[sheet_name_lower.find('увнк'):]: furnace_value = 'УВНК-9А2'
            else: furnace_value = 'УВНК-9А1'

        merged_ranges = list(sheet.merged_cells.ranges)
        large_cells = self.find_large_merged_cells(sheet, merged_ranges)
        chain_ranges = self.find_chain_ranges(sheet, merged_ranges)
        
        all_data = []
        plavka_counter = {}
        pouring_after_perebor_counter = 0
        perebor_detected_in_sheet = False
        
        for chain_idx, (start_row, end_row) in enumerate(chain_ranges):
            chain_height = end_row - start_row + 1
            if not has_uvnk:
                 pouring_after_perebor_counter = 0
                 perebor_detected_in_sheet = False
            
            first_block_start = 1 if has_uvnk else 2
            
            for block_idx, col_start in enumerate(range(first_block_start, sheet.max_column + 1, 8)):
                col_end = min(col_start + 7, sheet.max_column)
                if chain_height < 6: continue
                
                date_value = self.get_cell_value(sheet, merged_ranges, start_row, col_start)
                if not date_value: continue
                
                data_start_row = start_row + 3
                data_end_row = end_row - 3
                if data_start_row > data_end_row: continue
                
                if not has_uvnk:
                    furnace_value = self.find_furnace_for_non_uvnk(sheet, merged_ranges, start_row, col_start, col_end, data_start_row, data_end_row)
                    if furnace_value:
                        if "50" in str(furnace_value): furnace_value = "УППФ-50"
                        elif "У" in str(furnace_value).upper(): furnace_value = "УППФ-У"

                num_groups = (data_end_row - data_start_row + 1) // 3
                processed_metadata = set()

                for group_idx in range(num_groups):
                    group_start_row = data_start_row + group_idx * 3
                    if group_start_row + 2 > data_end_row: continue
                    
                    has_perebor = False
                    perebor_cell = None
                    if has_uvnk:
                        has_perebor, perebor_cell = self.check_group_for_perebor(group_start_row, col_start, large_cells)
                    
                    if has_perebor:
                        pouring_after_perebor_counter = 1
                        perebor_detected_in_sheet = True
                        self.log_message(f"    Группа {group_idx+1}: ОБНАРУЖЕНА ПЕРЕБОРКА!")
                        continue
                    
                    group_data = self.process_group(sheet, merged_ranges, group_start_row, col_start,
                                                date_value, furnace_value, has_uvnk, processed_metadata,
                                                sheet_name, chain_idx, block_idx, group_idx)
                    
                    if group_data:
                        if has_uvnk and perebor_detected_in_sheet:
                            group_data['Номер заливки после переборки'] = pouring_after_perebor_counter
                            pouring_after_perebor_counter += 1
                        
                        if group_data['Плавка']:
                            plavka_key = str(group_data['Плавка'])
                            plavka_counter[plavka_key] = plavka_counter.get(plavka_key, 0) + 1
                            group_data['Порядок заливки в плавке'] = plavka_counter[plavka_key]
                        
                        all_data.append(group_data)

        self.log_message(f"  Извлечено строк: {len(all_data)}")
        return all_data

    def process_directory(self, directory, output_file, update_progress=None):
        try:
            self.error_log = []
            
            excel_files = glob.glob(os.path.join(directory, "**", "*.xlsx"), recursive=True)
            excel_files.extend(glob.glob(os.path.join(directory, "**", "*.xls"), recursive=True))
            
            if not excel_files:
                self.log_message("Excel файлы не найдены")
                return
            
            self.log_message(f"Найдено {len(excel_files)} Excel файлов")
            
            all_data = []
            processed_files = 0
            
            for file_idx, input_file in enumerate(excel_files):
                if update_progress:
                     update_progress(int((file_idx) / len(excel_files) * 100))
                
                try:
                    file_name = os.path.basename(input_file)
                    self.log_message(f"[{file_idx+1}/{len(excel_files)}] Обработка: {file_name}")
                    
                    workbook = load_workbook(filename=input_file, data_only=True)
                    file_data = []
                    for sheet_name in workbook.sheetnames:
                        sheet_data = self.process_sheet(workbook, sheet_name)
                        file_data.extend(sheet_data)
                    
                    all_data.extend(file_data)
                    workbook.close()
                    processed_files += 1
                except Exception as e:
                    self.log_message(f"Ошибка с файлом {input_file}: {e}")
                    self.error_log.append({'Файл': input_file, 'Ошибка': str(e)})

            if update_progress: update_progress(100)
            
            if all_data:
                self.log_message("Формирование итоговой таблицы...")
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df = pd.DataFrame(all_data)
                    base_columns = ['Дата', 'Печь', 'Керамика', 'Железо', 'Плавка', 'ТУ',
                                'Металл', 'Шихта', 'Блок 1', 'Блок 2', 
                                'Комплект оснастки', 'Образцы', 'Номенклатура']
                    extra_columns = ['Номер заливки после переборки', 'Порядок заливки в плавке']
                    cols = base_columns + extra_columns + ['Комментарий']
                    
                    for col in cols:
                        if col not in df.columns: df[col] = None
                    
                    df[cols].to_excel(writer, sheet_name='Данные', index=False)
                    
                    if self.error_log:
                        pd.DataFrame(self.error_log).to_excel(writer, sheet_name='Ошибки', index=False)
                        
                    stats_data = {
                        'Показатель': ['Файлов', 'Строк', 'Ошибок', 'Дата'],
                        'Значение': [f"{processed_files}/{len(excel_files)}", len(all_data), 
                                   len(self.error_log), datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
                    }
                    pd.DataFrame(stats_data).to_excel(writer, sheet_name='Статистика', index=False)
                    
                self.log_message(f"Готово! Сохранено в {output_file}")
            else:
                self.log_message("Данные не найдены")

        except Exception as e:
            self.log_message(f"КРИТИЧЕСКАЯ ОШИБКА: {e}")
            traceback.print_exc()
