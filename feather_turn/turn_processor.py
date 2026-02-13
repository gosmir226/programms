import os
import pandas as pd
import json
import re
import hashlib
from datetime import datetime
from openpyxl import load_workbook
import traceback
import sys

# Добавляем родительскую директорию в путь для импорта общих модулей
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    from excel_manager import ExcelManager
    from cache_manager import CacheManager
except ImportError:
    ExcelManager = None
    CacheManager = None

class TurnProcessor:
    def __init__(self, directory, output_file=None, log_callback=None, progress_callback=None):
        self.directory = directory
        self.output_file = output_file or os.path.join(directory, 'consolidated_results.xlsx')
        self.log_callback = log_callback
        self.progress_callback = progress_callback

    def log_message(self, message):
        if self.log_callback:
            self.log_callback(message)
        else:
            print(message)

    def update_progress(self, value):
        if self.progress_callback:
            self.progress_callback(value)

    def clean_header(self, header):
        """Очищает заголовок от лишних пробелов"""
        if pd.isna(header):
            return ""
        header_str = str(header).strip()
        header_str = re.sub(r'\s+', ' ', header_str)
        return header_str

    def standardize_headers(self, headers):
        """Стандартизирует заголовки согласно mapping с обработкой дубликатов"""
        column_mapping = {
            'Плавка': ['Плавка'],
            'Отливка': ['Отливка', '№ отливки'],
            'Замер факт': ['Замер факт', 'Значение разворота по приспособлению ОДК-Климов', 'Значение разворота по приспособлению к4', 'Значение разворота по приспособлению к2'],
            'Статус': ['Статус4', 'Статус2', 'Статус'],
            'Контролер': ['Контролер', 'Исполнитель'],
            'Замер с поправкой': ['Замер с поправкой'],
            'Коэффициент': ['Коэффициент'],
            'Замер после доработки': ['Замер после доработки'],
            'Статус после доработки': ['Статус после доработки'],
            'после АТОС': ['после АТОС']
        }
        
        reverse_mapping = {}
        for standard_name, variants in column_mapping.items():
            for variant in variants:
                reverse_mapping[self.clean_header(variant)] = standard_name
        
        standardized_headers = []
        used_headers = set()
        
        for header in headers:
            clean_header = self.clean_header(header)
            
            if clean_header in reverse_mapping:
                standard_name = reverse_mapping[clean_header]
            else:
                standard_name = clean_header
            
            if standard_name in used_headers:
                counter = 1
                new_name = f"{standard_name}_{counter}"
                while new_name in used_headers:
                    counter += 1
                    new_name = f"{standard_name}_{counter}"
                standardized_headers.append(new_name)
                used_headers.add(new_name)
            else:
                standardized_headers.append(standard_name)
                used_headers.add(standard_name)
        
        return standardized_headers

    def process_coefficient_column(self, df, headers):
        """Обрабатывает столбец Коэффициент - берет значение из первой строки"""
        if 'Коэффициент' in headers:
            try:
                coeff_idx = list(headers).index('Коэффициент')
                if len(df) > 0:
                    first_value = df.iloc[0, coeff_idx]
                    if pd.notna(first_value):
                        df.iloc[:, coeff_idx] = first_value
            except (ValueError, IndexError):
                pass
        return df

    def extract_melt_info(self, melt_string):
        """Извлекает год и номер плавки из строки формата 25В11"""
        if pd.isna(melt_string):
            return None, None
        
        melt_str = str(melt_string).strip()
        match = re.search(r'(\d+)[ВB](\d+)', melt_str, re.IGNORECASE)
        if match:
            year = match.group(1)
            number = match.group(2)
            return year, number
        else:
            parts = re.split(r'[^\d]', melt_str, 1)
            if len(parts) >= 2:
                return parts[0], parts[1]
            elif len(parts) == 1:
                return parts[0], None
            else:
                return None, None

    def determine_status(self, value):
        """Определяет статус на основе числового значения"""
        try:
            num_value = float(value)
            if (-50 <= num_value <= -45) or (35 <= num_value <= 50):
                return "ОТРЫВ"
            elif (-40 <= num_value <= -30) or (20 <= num_value <= 30):
                return "ТР"
            elif -25 <= num_value <= 15:
                return "ГОД"
            else:
                return "БРАК"
        except (ValueError, TypeError):
            return "БРАК"

    def process_atos_rejection(self, df):
        """Обрабатывает столбцы для определения брака АТОС и обновления статуса"""
        standard_statuses = ["ГОД", "БРАК", "ТР", "ОТРЫВ", "ГОДН"]
        df['Брак АТОС'] = None
        
        atos_columns = ['Замер', 'Замер факт', 'Значение разворота по приспособлению ОДК-Климов', 
                       'Значение разворота по приспособлению к4', 'Значение разворота по приспособлению к2']
        correction_columns = ['Замер с поправкой']
        
        for idx, row in df.iterrows():
            current_status = str(row.get('Статус', '')).strip().upper() if pd.notna(row.get('Статус')) else ''
            contains_viz = 'ВИЗ' in current_status
            
            atos_detected = False
            for col in atos_columns:
                if col in df.columns and pd.notna(row.get(col)):
                    cell_value = str(row[col]).lower()
                    if 'атос' in cell_value or 'atos' in cell_value:
                        atos_detected = True
                        break
            
            if current_status not in standard_statuses and not contains_viz:
                measurement_value = None
                if atos_detected:
                    for col in correction_columns:
                        if col in df.columns and pd.notna(row.get(col)):
                            try:
                                measurement_value = float(row[col])
                                break
                            except (ValueError, TypeError):
                                continue
                else:
                    for col in correction_columns:
                        if col in df.columns and pd.notna(row.get(col)):
                            try:
                                measurement_value = float(row[col])
                                break
                            except (ValueError, TypeError):
                                continue
                    
                    if measurement_value is None:
                        for col in atos_columns:
                            if col in df.columns and pd.notna(row.get(col)):
                                try:
                                    measurement_value = float(row[col])
                                    break
                                except (ValueError, TypeError):
                                    continue
                
                if measurement_value is not None:
                    new_status = self.determine_status(measurement_value)
                    df.at[idx, 'Брак АТОС'] = "Брак АТОС"
                    df.at[idx, 'Статус'] = new_status
        
        return df

    def find_table_start(self, df):
        """Находит начальные координаты таблицы в DataFrame"""
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                if pd.notna(df.iloc[row_idx, col_idx]):
                    row_values = df.iloc[row_idx, col_idx:]
                    if row_values.notna().sum() >= 2:
                        return row_idx, col_idx
        return None, None

    def analyze_file_structure(self, file_path):
        """Анализирует структуру файла: количество листов и общее количество строк"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            sheets_count = len(workbook.sheetnames)
            total_rows = 0
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                row_count = 0
                for row in sheet.iter_rows():
                    if any(cell.value is not None for cell in row):
                        row_count += 1
                total_rows += row_count
            workbook.close()
            return sheets_count, total_rows
        except Exception:
            return 0, 0

    def extract_cassette_from_filename(self, file_path):
        filename = os.path.basename(file_path)
        match = re.search(r'№\s*(\d+)', filename)
        return match.group(1) if match else None

    def extract_cassette_from_sheetname(self, sheet_name):
        match = re.search(r'№\s*(\d+)', sheet_name)
        return match.group(1) if match else None

    def extract_cassette_and_measurement_info(self, file_path, sheet_name, file_type, measurement_rank=1):
        if file_type == 'первичный':
            cassette_number = self.extract_cassette_from_filename(file_path)
            measurement_number = 0
        else:
            cassette_number = self.extract_cassette_from_sheetname(sheet_name)
            measurement_number = measurement_rank
        return cassette_number, measurement_number

    def determine_file_types_in_folder(self, folder_path):
        """Определяет тип файлов в папке"""
        xlsx_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.xlsx')]
        file_types = {}
        
        primary_by_name = []
        remeasure_by_name = []
        undetermined = []
        
        for filename in xlsx_files:
            lower_name = filename.lower()
            if 'перезамер' in lower_name or 'доработк' in lower_name:
                remeasure_by_name.append(filename)
                file_types[filename] = 'перезамер'
            elif 'угол разворота' in lower_name:
                primary_by_name.append(filename)
                file_types[filename] = 'первичный'
            else:
                undetermined.append(filename)
        
        if undetermined:
            if len(undetermined) == 1:
                file_types[undetermined[0]] = 'первичный'
            else:
                file_stats = []
                for filename in undetermined:
                    file_path = os.path.join(folder_path, filename)
                    sheets_count, total_rows = self.analyze_file_structure(file_path)
                    file_stats.append({
                        'filename': filename,
                        'sheets_count': sheets_count,
                        'total_rows': total_rows
                    })
                
                file_stats.sort(key=lambda x: (x['sheets_count'], x['total_rows']), reverse=True)
                if not file_stats: return file_types

                max_sheets = max(stats['sheets_count'] for stats in file_stats)
                files_with_max_sheets = [stats for stats in file_stats if stats['sheets_count'] == max_sheets]
                
                if len(files_with_max_sheets) == 1:
                    file_types[files_with_max_sheets[0]['filename']] = 'перезамер'
                    for stats in file_stats:
                        if stats['filename'] != files_with_max_sheets[0]['filename']:
                             file_types[stats['filename']] = 'первичный'
                else:
                    files_with_max_sheets.sort(key=lambda x: x['total_rows'], reverse=True)
                    file_types[files_with_max_sheets[0]['filename']] = 'первичный'
                    for stats in files_with_max_sheets[1:]:
                        file_types[stats['filename']] = 'перезамер'
        
        return file_types

    def get_visible_sheets(self, file_path):
        """Возвращает список видимых листов в файле Excel"""
        try:
            wb = load_workbook(file_path, read_only=True)
            visible_sheets = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if sheet.sheet_state == 'visible':
                    visible_sheets.append(sheet_name)
            wb.close()
            return visible_sheets
        except Exception:
            return []

    def process_excel_file(self, file_path, file_type):
        """Обрабатывает один Excel-файл"""
        try:
            visible_sheets = self.get_visible_sheets(file_path)
            if not visible_sheets:
                return pd.DataFrame()
                
            excel_file = pd.ExcelFile(file_path)
            file_dataframes = []
            
            sheet_stats = []
            if file_type == 'перезамер':
                for sheet_name in visible_sheets:
                    if sheet_name not in excel_file.sheet_names: continue
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, engine='openpyxl')
                    start_row, start_col = self.find_table_start(df)
                    if start_row is None: continue
                    
                    table_df = df.iloc[start_row:, start_col:].copy()
                    table_df = table_df.reset_index(drop=True)
                    if len(table_df) > 0:
                        original_headers = [str(h) if pd.notna(h) else f"Unnamed_{i}" for i, h in enumerate(table_df.iloc[0])]
                        table_df.columns = original_headers
                        table_df = table_df.drop(0).reset_index(drop=True)
                    
                    if len(table_df) == 0: continue
                    if 'Плавка' in table_df.columns:
                        table_df = table_df[table_df['Плавка'].notna()]
                    
                    sheet_stats.append({'sheet_name': sheet_name, 'row_count': len(table_df), 'table_df': table_df})
                
                sheet_stats.sort(key=lambda x: x['row_count'], reverse=True)
                for rank, stat in enumerate(sheet_stats, 1):
                    df_proc = self.finalize_dataframe(stat['table_df'], file_path, stat['sheet_name'], file_type, rank)
                    if not df_proc.empty: file_dataframes.append(df_proc)
            else:
                for sheet_name in visible_sheets:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, engine='openpyxl')
                    start_row, start_col = self.find_table_start(df)
                    if start_row is None: continue
                    
                    table_df = df.iloc[start_row:, start_col:].copy()
                    table_df = table_df.reset_index(drop=True)
                    if len(table_df) > 0:
                        original_headers = [str(h) if pd.notna(h) else f"Unnamed_{i}" for i, h in enumerate(table_df.iloc[0])]
                        table_df.columns = original_headers
                        table_df = table_df.drop(0).reset_index(drop=True)
                    
                    if len(table_df) == 0: continue
                    if 'Плавка' in table_df.columns:
                        table_df = table_df[table_df['Плавка'].notna()]
                    
                    df_proc = self.finalize_dataframe(table_df, file_path, sheet_name, file_type)
                    if not df_proc.empty: file_dataframes.append(df_proc)
            
            if file_dataframes:
                return pd.concat(file_dataframes, ignore_index=True, sort=False)
            return pd.DataFrame()
        except Exception as e:
            self.log_message(f"Ошибка в process_excel_file {file_path}: {e}")
            return pd.DataFrame()

    def finalize_dataframe(self, table_df, file_path, sheet_name, file_type, rank=1):
        if table_df.empty: return table_df
        
        headers = self.standardize_headers(table_df.columns)
        table_df.columns = headers
        table_df = self.process_coefficient_column(table_df, headers)
        
        if 'Плавка' in table_df.columns:
            melt_info = table_df['Плавка'].apply(self.extract_melt_info)
            table_df['Год плавки'] = melt_info.apply(lambda x: x[0] if x else None)
            table_df['Номер плавки'] = melt_info.apply(lambda x: x[1] if x else None)
            table_df = table_df.rename(columns={'Плавка': 'Полный номер плавки'})
        
        table_df = self.process_atos_rejection(table_df)
        cassette, measurement = self.extract_cassette_and_measurement_info(file_path, sheet_name, file_type, rank)
        
        table_df['Имя файла'] = os.path.basename(file_path)
        table_df['Номер замера'] = measurement
        table_df['Номер кассеты'] = cassette
        table_df['Тип файла'] = 'Перезамер' if file_type == 'перезамер' else 'Первичный'
        table_df['Полный путь'] = file_path
        
        return table_df

    def process(self):
        try:
            result_path = self.output_file
            
            if CacheManager:
                cache = CacheManager("feather_turn", result_path)
            else:
                cache = None
                
            if ExcelManager:
                excel_manager = ExcelManager(result_path)
            else:
                excel_manager = None

            new_data_frames = []
            folders = [f for f in os.listdir(self.directory) 
                      if os.path.isdir(os.path.join(self.directory, f)) and f != '__pycache__']
            
            # Если в самой директории есть файлы, добавим и её для анализа (некоторые пользователи кладут файлы прямо туда)
            folders = ['.'] + folders
            
            total_folders = len(folders)
            processed_files_count = 0
            skipped_files_count = 0
            
            for folder_idx, folder in enumerate(folders):
                folder_path = os.path.join(self.directory, folder) if folder != '.' else self.directory
                
                xlsx_files = [f for f in os.listdir(folder_path) 
                             if f.lower().endswith('.xlsx') and not f.startswith('~$')]
                
                # Исключаем выходной файл
                xlsx_files = [f for f in xlsx_files if os.path.abspath(os.path.join(folder_path, f)) != os.path.abspath(result_path)]
                
                if not xlsx_files: continue
                
                self.log_message(f"Анализ папки: {folder if folder != '.' else 'Корень'}")
                file_types = self.determine_file_types_in_folder(folder_path)
                
                xlsx_files_with_paths = [(f, os.path.join(folder_path, f)) for f in xlsx_files]
                xlsx_files_with_paths.sort(key=lambda x: os.path.getmtime(x[1]))
                
                for filename, file_path in xlsx_files_with_paths:
                    file_type = file_types.get(filename, 'первичный')
                    abs_file_path = os.path.abspath(file_path)
                    
                    if cache and not cache.is_file_changed(abs_file_path):
                        skipped_files_count += 1
                        continue
                    
                    self.log_message(f"  Обработка ({file_type}): {filename}")
                    file_data = self.process_excel_file(file_path, file_type)
                    
                    if not file_data.empty:
                        new_data_frames.append(file_data)
                        if cache:
                            cache.update_file(abs_file_path)
                        processed_files_count += 1
                    else:
                        if cache: cache.update_file(abs_file_path)
                
                self.update_progress(int((folder_idx + 1) / total_folders * 100))
            
            if new_data_frames:
                self.log_message("Объединение новых данных...")
                final_new_data = pd.concat(new_data_frames, ignore_index=True, sort=False)
                
                if excel_manager:
                    self.log_message("Сохранение в Excel...")
                    excel_manager.write_excel_smart(
                        final_new_data,
                        key_columns=['Полный номер плавки', 'Отливка', 'Номер кассеты', 'Номер замера'],
                        sheet_name='Sheet1',
                        log_callback=self.log_message
                    )
                    final_df = excel_manager.read_excel_smart('Sheet1')[0]
                else:
                    final_df = final_new_data
                    final_df.to_excel(result_path, index=False, engine='openpyxl')
                
                self.log_message(f"Готово! Обработано новых: {processed_files_count}, пропущено: {skipped_files_count}")
                return final_df
            else:
                self.log_message(f"Новых данных не обнаружено. Пропущено: {skipped_files_count}")
                if excel_manager:
                    return excel_manager.read_excel_smart('Sheet1')[0]
                return pd.DataFrame()
            
        except Exception as e:
            self.log_message(f"Ошибка в process: {str(e)}")
            self.log_message(traceback.format_exc())
            return pd.DataFrame()

def clear_cache_for_output(result_path):
    if CacheManager:
        cache = CacheManager("feather_turn", result_path)
        cache.clear_cache()
