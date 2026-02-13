"""
Модуль для надежного управления Excel файлами.

Предоставляет функции для умного чтения и записи Excel файлов с сохранением
пользовательских изменений (новые столбцы, строки, порядок столбцов).
"""

import pandas as pd
import os
from typing import List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelManager:
    """Управление Excel файлами с сохранением пользовательских изменений."""
    
    def __init__(self, file_path: str):
        """
        Инициализация менеджера Excel файла.
        
        Args:
            file_path: Путь к Excel файлу
        """
        self.file_path = file_path
        self.exists = os.path.exists(file_path)
    
    def read_excel_smart(self, sheet_name: str = 0) -> Tuple[pd.DataFrame, dict]:
        """
        Умное чтение Excel с сохранением метаданных.
        
        Args:
            sheet_name: Название или индекс листа
            
        Returns:
            Tuple[DataFrame, dict]: Данные и метаданные (порядок столбцов, доп. столбцы)
        """
        if not self.exists:
            return pd.DataFrame(), {'column_order': [], 'user_columns': []}
        
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine='openpyxl')
            
            metadata = {
                'column_order': list(df.columns),
                'user_columns': [],  # Будет определено при обновлении
                'original_row_count': len(df)
            }
            
            return df, metadata
            
        except Exception as e:
            print(f"Ошибка при чтении Excel файла {self.file_path}: {e}")
            return pd.DataFrame(), {'column_order': [], 'user_columns': []}

    def read_existing_data(self, sheet_name: str = 0) -> pd.DataFrame:
        """
        Упрощенное чтение существующих данных.
        
        Args:
            sheet_name: Название или индекс листа
            
        Returns:
            pd.DataFrame: DataFrame с данными или пустой DataFrame
        """
        df, _ = self.read_excel_smart(sheet_name)
        return df
    
    def write_excel_smart(
        self, 
        new_data: pd.DataFrame, 
        key_columns: List[str],
        sheet_name: str = 'Data',
        log_callback=None,
        mode: str = 'update',
        format_as_table: bool = True
    ) -> bool:
        """
        Умная запись данных в Excel с сохранением пользовательских изменений.
        
        Args:
            new_data: Новые данные для записи
            key_columns: Столбцы для идентификации строк (ключи)
            sheet_name: Название листа
            log_callback: Функция для логирования
            mode: Режим записи ('update', 'append', 'replace')
            format_as_table: Форматировать ли данные как таблицу Excel
            
        Returns:
            bool: Успешность операции
        """
        def log(msg):
            if log_callback:
                log_callback(msg)
            else:
                print(msg)
        
        try:
            # Читаем существующие данные, если не режим replace
            if self.exists and mode != 'replace':
                existing_data, metadata = self.read_excel_smart(sheet_name)
            else:
                existing_data = pd.DataFrame()
                metadata = {'column_order': [], 'user_columns': []}

            if not existing_data.empty:
                log(f"⚠️ Файл {os.path.basename(self.file_path)} существует. Выполняется умное обновление данных...")
                
                # Определяем пользовательские столбцы
                user_columns = [col for col in existing_data.columns if col not in new_data.columns]
                if user_columns:
                    log(f"Обнаружены пользовательские столбцы: {', '.join(user_columns)}")
                
                # Проверяем ключи
                missing_keys = [key for key in key_columns if key not in new_data.columns]
                if missing_keys and mode == 'update':
                    log(f"⚠️ Ключевые столбцы {missing_keys} отсутствуют в новых данных. Переключение в append.")
                    mode = 'append'
                
                if mode == 'append':
                    combined = pd.concat([existing_data, new_data], ignore_index=True, sort=False)
                    log(f"Добавлено {len(new_data)} новых строк")
                else:
                    combined = self._merge_data_smart(existing_data, new_data, key_columns, user_columns, log)
                
                # Порядок столбцов
                original_columns = metadata['column_order']
                new_cols = [col for col in combined.columns if col not in original_columns]
                final_column_order = [col for col in original_columns if col in combined.columns] + new_cols
                combined = combined[final_column_order]
            else:
                combined = new_data
                if self.exists and mode == 'replace':
                    log(f"Заменяем файл новым: {self.file_path}")
                else:
                    log(f"Создаем новый файл: {self.file_path}")

            # Записываем
            writer_args = {'engine': 'openpyxl'}
            if self.exists and mode != 'replace':
                writer_args['mode'] = 'a'
                writer_args['if_sheet_exists'] = 'replace'
            else:
                writer_args['mode'] = 'w'

            with pd.ExcelWriter(self.file_path, **writer_args) as writer:
                combined.to_excel(writer, sheet_name=sheet_name, index=False)
                
                if format_as_table and not combined.empty:
                    from openpyxl.worksheet.table import Table, TableStyleInfo
                    from openpyxl.utils import get_column_letter
                    
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # 1. Форматируем как таблицу
                    row_count = len(combined)
                    col_count = len(combined.columns)
                    ref = f"A1:{get_column_letter(col_count)}{row_count + 1}"
                    
                    # Удаляем старые таблицы
                    for table in list(worksheet.tables.values()):
                        del worksheet.tables[table.name]
                        
                    table_name = "".join(filter(str.isalnum, sheet_name)) or "Data"
                    table = Table(displayName=f"Table_{table_name}", ref=ref)
                    style = TableStyleInfo(
                        name="TableStyleMedium2",
                        showFirstColumn=False, showLastColumn=False,
                        showRowStripes=True, showColumnStripes=False
                    )
                    table.tableStyleInfo = style
                    worksheet.add_table(table)
                    
                    # 2. Автоширина столбцов
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except: pass
                        adjusted_width = min(max_length + 2, 60)
                        worksheet.column_dimensions[column_letter].width = adjusted_width

            self.exists = True
            log(f"✅ Данные сохранены в {os.path.basename(self.file_path)}. Строк: {len(combined)}")
            return True
            
        except Exception as e:
            log(f"❌ Ошибка записи Excel: {e}")
            import traceback
            log(traceback.format_exc())
            return False

    def _merge_data_smart(
        self, 
        existing: pd.DataFrame, 
        new: pd.DataFrame, 
        key_columns: List[str],
        user_columns: List[str],
        log_callback
    ) -> pd.DataFrame:
        """
        Объединение существующих и новых данных с сохранением пользовательских изменений.
        """
        # Создаем составной ключ
        def make_key(df, cols):
            return df[cols].astype(str).agg('|'.join, axis=1)
        
        try:
            existing_keys = make_key(existing, key_columns)
            new_keys = make_key(new, key_columns)
        except KeyError as e:
            log_callback(f"⚠️ Ключевой столбец отсутствует: {e}. Используется режим append.")
            return pd.concat([existing, new], ignore_index=True, sort=False)
        
        # Индексы
        rows_to_update = existing_keys.isin(new_keys)
        rows_to_keep = ~rows_to_update
        
        log_callback(f"   Обновляется (перезаписывается): {rows_to_update.sum()}")
        log_callback(f"   Сохраняется без изменений: {rows_to_keep.sum()}")
        log_callback(f"   Добавляется новых: {(~new_keys.isin(existing_keys)).sum()}")
        
        # Сохранение пользовательских данных для обновляемых строк
        if user_columns and rows_to_update.any():
            user_data = existing.loc[rows_to_update, key_columns + user_columns].copy()
            # Убираем дубликаты в user_data по ключам (на всякий случай)
            user_data = user_data.drop_duplicates(subset=key_columns)
            
            new = new.merge(user_data, on=key_columns, how='left', suffixes=('', '_user'))
            
            for col in user_columns:
                if col not in new.columns and f"{col}_user" in new.columns:
                    new[col] = new[f"{col}_user"]
                if f"{col}_user" in new.columns:
                    new.drop(columns=[f"{col}_user"], inplace=True)
        
        unchanged_rows = existing[rows_to_keep].copy()
        return pd.concat([unchanged_rows, new], ignore_index=True, sort=False)


def write_excel_multiple_sheets(
    file_path: str,
    sheets_data: dict,
    log_callback=None,
    format_as_table: bool = True
) -> bool:
    """
    Записать несколько листов в Excel файл.
    """
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)
    
    try:
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, df in sheets_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                if format_as_table and not df.empty:
                    worksheet = writer.sheets[sheet_name]
                    row_count = len(df)
                    col_count = len(df.columns)
                    ref = f"A1:{get_column_letter(col_count)}{row_count + 1}"
                    
                    table_name = "".join(filter(str.isalnum, sheet_name)) or "Sheet"
                    table = Table(displayName=f"Table_{table_name}", ref=ref)
                    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
                    worksheet.add_table(table)
                    
                    for column in worksheet.columns:
                        max_length = max([len(str(cell.value) or "") for cell in column])
                        worksheet.column_dimensions[get_column_letter(column[0].column)].width = min(max_length + 2, 60)
        
        log(f"✅ Записано {len(sheets_data)} листов в файл {os.path.basename(file_path)}")
        return True
    except Exception as e:
        log(f"❌ Ошибка при записи файла: {e}")
        return False
